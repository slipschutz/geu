#!/usr/bin/env python3


import sys
sys.path.append("xlwt-0.7.5")

sys.path.append('./.local/lib/python3.4/site-packages/')


from openpyxl import Workbook

#import xlwt

from CBULine import *
from DeductionLine import *

import CBULoader
import DeductionLoader

import ReconciledEntry 


from os import listdir
from os.path import isfile, join
import os
import datetime

def AttemptBlankNetIdRecovery(CBUWrap,DeductionWrap): #EmptyNetIdMap,DuesFirstLastMap,GoodCBUNetIdList):
    print("Don't use this function ")

        
    #loop over all of the empty net ids in the CBU list
    #That map should be from firstLast names to info
    for key, info in CBUWrap.EmptyNetIdMap.items():
        #First Try to find based on first name and lastname from the DUES list
        if key in DeductionWrap.FirstLastNameMap:
            #The first last name key is in the dedcutions map
            #the record can be recovered
            print ("recored recovered")
            #info.NetId=irstLastMap[key].NetId
            #GoodCBUNetIdList[DuesFirstLastMap[key].NetId]=info
            
        if info.PersonnelNumber in DeductionWrap.EmployeeNumberMap:
            print ("victory")
#            netid=DuesDeductionEmployeeIDs[info.PersonnelNumber].Lines[0].NetId
 #           info.NetId=netid
  #          GoodCBUNetIdList[netid]=info



            
def AttemptCBUEmployeeNumberRecovery(CBUNetId,DeductionNetId):
    for netid, CBUInfo in CBUNetId.iteritems():
        if CBUInfo.PersonnelNumber==" ":
            print ("DDDD")
            

def Reconcile(CBU_File,Dues_File):

 
    CBUWrapper=  CBULoader.LoadCBU(CBU_File) 
    DeductionsWrapper= DeductionLoader.LoadDeduction(Dues_File)


    Dues_File_Temp=os.path.basename(Dues_File)
    splitName=Dues_File_Temp.split("-")
    datePart=splitName[0].strip()
    dateTemp=datetime.datetime.strptime(datePart,"%Y%m%d").date()
    DefualtDate=str(dateTemp.strftime("%m/%d/%Y"))
    

    MapForInCBUButNotInDeductions={}
    ReconciledMap={}

    


    MapOfAllNetIds={}
    
    DuesID=DeductionsWrapper.NetIdMap
    CBUID=CBUWrapper.NetIdMap

    
    for netid,info in CBUID.items():
        if "fees" in info.DuesWageType.lower().strip():
            info.DuesWageType="GEU Fees-C"
        if "dues" in info.DuesWageType.lower().strip():
            info.DuesWageType="GEU Dues"
        MapOfAllNetIds[netid]=info

    for netid,info in DuesID.items():
        for line in info.Lines:
            if "fees" in line.WageTypeText.lower().strip():
                line.WageTypeText="GEU Fees-C"
            if "dues" in line.WageTypeText.lower().strip():
                line.WageTypeText="GEU Dues"
        MapOfAllNetIds[netid]=info


    for netid, info in MapOfAllNetIds.items():
        if netid in DuesID and netid in CBUID:
            ##It is in both lists
            CBUInfo =CBUID[netid]
            DuesInfo=DuesID[netid]
            temp=ReconciledEntry.ReconciledEntry()
            temp.CopyCBUInfo(CBUInfo)
            temp.CopyDuesInfo(DuesInfo)

            if CBUInfo.DuesWageType.lower() != DuesInfo.Lines[0].WageTypeText.lower():

                ## CBU Entry needs updating
                temp.SetValueByTag("UpdatedWageType",DuesInfo.Lines[0].WageTypeText)
                temp.SetValueByTag("WasUpdated","yes")

            else: # does not need updating in the CBU list

                temp.SetValueByTag("UpdatedWageType",CBUInfo.DuesWageType)
                temp.SetValueByTag("WasUpdated","no")
                # CBULine.UpdatedDuesType =CBULine.DuesType
                # CBULine.WasUpdated="no"
            #now put entry in the MapFor Both Lists 
            ReconciledMap[netid]=temp
        elif netid in DuesID and netid not in CBUID:

            DuesInfo=DuesID[netid]
            temp=ReconciledEntry.ReconciledEntry()
            temp.CopyDuesInfo(DuesInfo)
            temp.SetValueByTag("MSUNETID",netid.upper())
            temp.SetValueByTag("OnlyInDues","Yes")
            ReconciledMap[netid]=temp

        elif netid in CBUID and netid not in DuesID:
            CBUInfo =CBUID[netid]
            temp=ReconciledEntry.ReconciledEntry()
            temp.CopyCBUInfo(CBUInfo)
            temp.SetValueByTag("UpdatedWageType","Blank")
            temp.SetValueByTag("OnlyInCBU","Yes")
            ReconciledMap[netid]=temp

    
    print ("length is ",len(ReconciledMap))
    wb = Workbook()

    temp =os.path.split(CBU_File)[1]
    temp2=os.path.split(Dues_File)[1]
    
    temp3 = temp.split('.')
    temp4 = temp2.split('.')


    dest_filename =join("ReconciledData",temp3[0]+"_"+temp4[0]+".xlsx")

    ws0 = wb.active
    ws0.title = "Reconciled List"


    count=1
    for i in ReconciledEntry.ListOfKnackNames:
        ws0.cell(row=1,column=count,value=i)
        count=count+1


    count=2
    for netid,line in ReconciledMap.items():
            
        if len(line.DuesPersonInfo.Lines) == 0:
            line.SetValueByTag("Date",DefualtDate)
            for i in range(len(ReconciledEntry.ListOfColumnNames)):
                ws0.cell(row=count, column=i+1,value=line.GetValueByIndex(i))
            ws0.cell(row=count,column=len(ReconciledEntry.ListOfColumnNames)+1,value="Yes")
            count=count+1
        else:
            for j in range(len(line.DuesPersonInfo.Lines)):
                line.SetDuesPersonInfo(j)#Choose one of the entries from the deduction file
                for i in range(len(ReconciledEntry.ListOfColumnNames)):
                    ws0.cell(row=count, column=i+1,value=line.GetValueByIndex(i))
                ws0.cell(row=count,column=len(ReconciledEntry.ListOfColumnNames)+1,value="Yes")
                count=count+1

    print(dest_filename)
    wb.save(filename = dest_filename)
    
    return dest_filename



    
    for NetId,CBULine in MapForCBU.iteritems():
        if NetId in MapForDeductionsList:         #NetId is in both lists
            #first copy over info from CBULine to ReconciledEntry object
            temp=ReconciledEntry.ReconciledEntry()
            temp.CopyCBUInfo(CBULine)
            temp.CopyDuesInfo(MapForDeductionsList[NetId])
            if CBULine.DuesType.lower() != MapForDeductionsList[NetId].Lines[0].WageTypeText.lower():
                ## CBU Entry needs updating
                temp.SetValueByTag("UpdatedWageType",MapForDeductionsList[NetId].Lines[0].WageTypeText)
                temp.SetValueByTag("WasUpdated","yes")

            else: # does not need updating in the CBU list 
                temp.SetValueByTag("UpdatedWageType",CBULine.DuesType)
                temp.SetValueByTag("WasUpdated","no")
                # CBULine.UpdatedDuesType =CBULine.DuesType
                # CBULine.WasUpdated="no"
            #now put entry in the MapFor Both Lists 
            ReconciledMap[NetId]=temp

        #Entry is in the CBU list but does not have a matching net ID in deductions
        #Look to see if this CBULine matchs based on last and first name
        else:
            tempkey="fakefakefake"
            if CBULine.FirstName != " " and CBULine.LastName!=" ":
                tempkey = (str(CBULine.FirstName)+str(CBULine.LastName)).strip().lower()

            if tempkey in FirstNameLastNameList:
                #This CBU line is in the deductions file but with wrong NETID    
                temp=ReconciledEntry.ReconciledEntry()
                temp.CopyCBUInfo(CBULine)
                duesNetId=FirstNameLastNameList[tempkey].Lines[0].NetId
                temp.CopyDuesInfo(FirstNameLastNameList[tempkey])
                temp.SetValueByTag("DuesFileNetId",duesNetId)

                if CBULine.DuesType.lower() != FirstNameLastNameList[tempkey].Lines[0].WageTypeText.lower():
                    temp.SetValueByTag("UpdatedWageType",\
                                           FirstNameLastNameList[tempkey].Lines[0].WageTypeText)
                    temp.SetValueByTag("WasUpdated","yes")
                else: # does not need updating in the CBU list 
                    temp.SetValueByTag("UpdatedWageType",CBULine.DuesType)
                    temp.SetValueByTag("WasUpdated","no")

                #Put the reconciled entry in to the map using the 
                #net id from the CBU list
                #which is still going to be unique
                ReconciledMap[NetId]=temp
                
            else:
                #Entry is really only in the CBU list
                MapForInCBUButNotInDeductions[NetId]=CBULine
    endfor=0

    print ("Recondiled SIZE IS ",len(ReconciledMap))

    for netid, line in MapForInCBUButNotInDeductions.iteritems():
        temp=ReconciledEntry.ReconciledEntry()
        temp.CopyCBUInfo(line)
        temp.SetValueByTag("OnlyInCBU","yes")
        temp.SetValueByTag("UpdatedWageType","Blank")#line.DuesType)
        temp.SetValueByTag("WasUpdated","no")
        ReconciledMap[netid]=temp

    print ("SIZE IS ",len(MapForInCBUButNotInDeductions))


### Do search from the dues list.  Look to see if there are things in here 
### That are not in the CBU List
    for netid, deductionLine in MapForDeductionsList.iteritems():
        if netid not in MapForCBU:
            tempKey=str(deductionLine.Lines[0].FirstName+deductionLine.Lines[0].LastName).lower()
        
            if tempKey not in MapForFirstLastCBU:
            ###There is a line in the deduction list that does
            #not match the CBU list by net id and does not match by first/last name
        
                temp=ReconciledEntry.ReconciledEntry()
                
                temp.CopyDuesInfo(deductionLine)
                temp.SetValueByTag("OnlyInDues","yes")
                temp.SetValueByTag("UpdatedWageType",deductionLine.Lines[0].WageTypeText)
                temp.SetValueByTag("WasUpdated","no")
                temp.SetValueByTag("MSUNETID",deductionLine.Lines[0].NetId)
                temp.SetValueByTag("DuesFileNetId",deductionLine.Lines[0].NetId)
                
                ReconciledMap[netid]=temp
        
            else:
                #this is a entry that does not match by net id but DOES match on last/first
                #these entries were already taken care of
                pass
        
            endif=0


    #Look if the CBU lines with no NetId match a first/last name in deduction list
    for netid,line in MapForEmptyNetIdCBU.iteritems():
        tempNameThing=line.FirstName+line.LastName
        tempNameThing=tempNameThing.lower()
        if tempNameThing in FirstNameLastNameList:
            print ("Found")

    
    Department2Members={}
    Department2FeePayers={}
    Department2NonComp={}

    for netid, line in ReconciledMap.iteritems():
        key =line.GetValueByTag("EnrolledUnitName").lower()

        if key in Department2Members:
            pass
        else:
            Department2Members[key]=0
            Department2FeePayers[key]=0
            Department2NonComp[key]=0

        if line.GetValueByTag("UpdatedWageType").lower() == "geu dues":
            Department2Members[key] +=1
        elif line.GetValueByTag("UpdatedWageType").lower() == "geu fees-c":
            Department2FeePayers[key] +=1
        elif line.GetValueByTag("UpdatedWageType").lower() == "geu fees-nc":
            Department2NonComp[key] +=1
        else:
            Department2NonComp[key] +=1

            
    for netid, line in ReconciledMap.iteritems():
        for j in line.DuesPersonInfo.Lines:
            if j.WageTypeText.lower().strip() == "geu dues":
                j.WageTypeText = "GEU Dues"
            if j.WageTypeText.lower().strip() == "geu fees-c":
                j.WageTypeText = "GEU Fees-C"
            
#########################Write everything to a file#######################
    wb = Workbook()
    ws0 = wb.add_sheet('ReconciledCBU')
    
    count=0 
    for i in ReconciledEntry.ListOfKnackNames:
        ws0.write(0,count,i)
        count=count+1

    count=1
    for netid,line in ReconciledMap.iteritems():
        if len(line.DuesPersonInfo.Lines) == 0:
            line.SetValueByTag("Date",DefualtDate)
            for i in range(len(ReconciledEntry.ListOfColumnNames)):
                ws0.write(count, i,line.GetValueByIndex(i))
            ws0.write(count,len(ReconciledEntry.ListOfColumnNames)+1,"Payroll Sheet Update")

            count=count+1
        else:
            for j in range(len(line.DuesPersonInfo.Lines)):
                line.SetDuesPersonInfo(j)#Choose one of the entries from the deduction file
                for i in range(len(ReconciledEntry.ListOfColumnNames)):
                    ws0.write(count, i,line.GetValueByIndex(i))
                ws0.write(count,len(ReconciledEntry.ListOfColumnNames)+1,"Payroll Sheet Update")
                count=count+1


    ws1 = wb.add_sheet('BlankNetIdInCBU')
    count=0
    for netid,line in MapForEmptyNetIdCBU.iteritems():
        for i in range(32):
            ws1.write(count, i,line.GetValue(i))
        count=count+1


    # ws2 = wb.add_sheet('CBUNotDeductWithDuesStatus')
    # count=0
    # for netid,line in MapForInCBUButNotInDeductionsWithDuesStatus.iteritems():
    #     for i in range(32):
    #         ws2.write(count, i,line.GetValue(i))
    #     count=count+1

    # ws3 = wb.add_sheet('CBUNotDeductWithNoDuesStatus')
    # count=0
    # for netid,line in MapForInCBUButNotInDeductionsWithNoDuesStatus.iteritems():
    #     for i in range(32):
    #         ws3.write(count, i,line.GetValue(i))
    #     count=count+1
    

   # print len(Department2Members)," ",len(Department2FeePayers)," ",len(Department2NonComp)
    
    ws4 = wb.add_sheet('DepartStats')
    count=1
    ws4.write(0,0,"Department")
    ws4.write(0,1,"Members")
    ws4.write(0,2,"FeePayers")
    ws4.write(0,3,"Non-Compliant")
    ws4.write(0,4,"Total members")
    totalNumberOfPeople=0
    totalMembersEveryWhere=0
    for name,val in Department2Members.iteritems():
        tot=float(Department2Members[name]+Department2FeePayers[name]+Department2NonComp[name])
        totalNumberOfPeople+=tot
        totalMembersEveryWhere+=float(Department2Members[name])
        ws4.write(count,0,name)
        if tot != 0 :
            ws4.write(count,1,(Department2Members[name]/tot) * 100)
            ws4.write(count,2,Department2FeePayers[name]/tot *100)
            ws4.write(count,3,Department2NonComp[name]/tot *100)
            ws4.write(count,4,tot )
        
        count=count+1

    ws4.write(count+1,1,totalMembersEveryWhere/totalNumberOfPeople)
    temp =os.path.split(CBU_File)[1]
    temp2=os.path.split(Dues_File)[1]
    
    temp3 = temp.split('.')
    temp4 = temp2.split('.')


    outFileName=join("ReconciledData",temp3[0]+"_"+temp4[0]+".xls")


    try: 
        wb.save(outFileName)
        #GuiWindow.DisplayMessageWindow("Created Reconciled File ")
    except IOError:
        print ("You probabaly need to close the OutPut File")

        

